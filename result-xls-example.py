#!/usr/bin/env python3

import datetime
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment


#result = {'run_number': 'nitin', 'flowcell_id': 'HMVWNBGX7', 'rta_version': '2.4.11', 'lanes': {'2': {'reads': [{'percent_pass_filter': 91.86, 'error_rate_stddev': 0.29, 'error_rate': 0.37, 'num_reads': 112923226, 'percent_q30': 89.39, 'num_cycles': 144, 'percent_aligned': 6.56, 'read_number': 1}, {'percent_pass_filter': 91.86, 'error_rate_stddev': 0, 'error_rate': 0, 'num_reads': 112923226, 'percent_q30': 96.36, 'num_cycles': 8, 'percent_aligned': 0, 'read_number': 2}, {'percent_pass_filter': 91.86, 'error_rate_stddev': 0, 'error_rate': 0, 'num_reads': 112923226, 'percent_q30': 95.02, 'num_cycles': 16, 'percent_aligned': 0, 'read_number': 3}, {'percent_pass_filter': 91.86, 'error_rate_stddev': 0.25, 'error_rate': 0.42, 'num_reads': 112923226, 'percent_q30': 88.26, 'num_cycles': 144, 'percent_aligned': 0, 'read_number': 4}], 'cluster_density': 189.52, 'cluster_density_stddev': 6.26}, '4': {'reads': [{'percent_pass_filter': 91.19, 'error_rate_stddev': 0.34, 'error_rate': 0.38, 'num_reads': 112884478, 'percent_q30': 89.02, 'num_cycles': 144, 'percent_aligned': 6.54, 'read_number': 1}, {'percent_pass_filter': 91.19, 'error_rate_stddev': 0, 'error_rate': 0, 'num_reads': 112884478, 'percent_q30': 96.31, 'num_cycles': 8, 'percent_aligned': 0, 'read_number': 2}, {'percent_pass_filter': 91.19, 'error_rate_stddev': 0, 'error_rate': 0, 'num_reads': 112884478, 'percent_q30': 94.96, 'num_cycles': 16, 'percent_aligned': 0, 'read_number': 3}, {'percent_pass_filter': 91.19, 'error_rate_stddev': 0.25, 'error_rate': 0.42, 'num_reads': 112884478, 'percent_q30': 88.08, 'num_cycles': 144, 'percent_aligned': 0, 'read_number': 4}], 'cluster_density': 190.83, 'cluster_density_stddev': 5.54}, '3': {'reads': [{'percent_pass_filter': 90.92, 'error_rate_stddev': 1.17, 'error_rate': 0.67, 'num_reads': 113822545, 'percent_q30': 86.55, 'num_cycles': 144, 'percent_aligned': 6.56, 'read_number': 1}, {'percent_pass_filter': 90.92, 'error_rate_stddev': 0, 'error_rate': 0, 'num_reads': 113822545, 'percent_q30': 96.64, 'num_cycles': 8, 'percent_aligned': 0, 'read_number': 2}, {'percent_pass_filter': 90.92, 'error_rate_stddev': 0, 'error_rate': 0, 'num_reads': 113822545, 'percent_q30': 94.88, 'num_cycles': 16, 'percent_aligned': 0, 'read_number': 3}, {'percent_pass_filter': 90.92, 'error_rate_stddev': 0.31, 'error_rate': 0.46, 'num_reads': 113822545, 'percent_q30': 86.87, 'num_cycles': 144, 'percent_aligned': 0, 'read_number': 4}], 'cluster_density': 192.99, 'cluster_density_stddev': 7.15}, '1': {'reads': [{'percent_pass_filter': 91.99, 'error_rate_stddev': 0.54, 'error_rate': 0.44, 'num_reads': 111523447, 'percent_q30': 88.47, 'num_cycles': 144, 'percent_aligned': 6.53, 'read_number': 1}, {'percent_pass_filter': 91.99, 'error_rate_stddev': 0, 'error_rate': 0, 'num_reads': 111523447, 'percent_q30': 96.86, 'num_cycles': 8, 'percent_aligned': 0, 'read_number': 2}, {'percent_pass_filter': 91.99, 'error_rate_stddev': 0, 'error_rate': 0, 'num_reads': 111523447, 'percent_q30': 95.17, 'num_cycles': 16, 'percent_aligned': 0, 'read_number': 3}, {'percent_pass_filter': 91.99, 'error_rate_stddev': 0.32, 'error_rate': 0.41, 'num_reads': 111523447, 'percent_q30': 88.33, 'num_cycles': 144, 'percent_aligned': 0, 'read_number': 4}], 'cluster_density': 186.9, 'cluster_density_stddev': 8.5}}, 'folder_location': '/scratch/cgssb/nextseq2/181015_NB501065_0286_AHMVWNBGX7', 'read_cycles': 288, 'total_cycles': 312, 'read_num_lookup': [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 2, 2, 2, 2, 2, 2, 2, 2, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4], 'instrument': 'NB501065', 'run_id': '181015_NB501065_0286_AHMVWNBGX7', 'index_cycles': 24}

wb = Workbook()
ws = wb.active
#ws['A1'] = 42
#ws.append([1, 2, 3])
#ws['A2'] = datetime.datetime.now()

columns = ("Density","Cluster PF","R1: Phasing","R1: Prephasing","Reads PF","R1: Q30","R1: Aligned","R1: ER","R1: ER (35)")

ws.append(columns)

rows = ws['A1:I1']

for row in rows:
    for c in row:
        c.fill = PatternFill('solid', fgColor="FDEADA")


wb.save("sample.xlsx")





